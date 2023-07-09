import openpyxl
import matplotlib.pyplot as plt
from datetime import date

def create_table(paragraph_file, attribute_file, output_name):
    # Read paragraphs from the first file
    with open(paragraph_file, 'r') as f:
        paragraphs = f.readlines()

    # Read attributes from the second file
    with open(attribute_file, 'r') as f:
        attributes = [attr.strip() for attr in f.readlines()]

    # Create a new workbook and select the active sheet for the table
    workbook = openpyxl.Workbook()
    table_sheet = workbook.active
    table_sheet.title = 'Table'

    # Write the attributes as the first row in the table sheet
    table_sheet.append([''] + attributes)

    # Process each paragraph and extract attribute values
    for i, paragraph in enumerate(paragraphs):
        attr_values = {}
        for attr in attributes:
            attr_key = attr + ':'
            if attr_key in paragraph:
                start_index = paragraph.index(attr_key) + len(attr_key)
                end_index = paragraph.find(' ', start_index)
                if end_index == -1:
                    end_index = None
                attr_value = paragraph[start_index:end_index].strip()
                attr_values[attr] = attr_value

        # Write attribute values as a new row in the table sheet
        row_values = [paragraph.strip()] + [attr_values.get(attr, '') for attr in attributes]
        table_sheet.append(row_values)

    # Save the workbook to the Excel file
    excel_output_file = f"{output_name}_{date.today().strftime('%Y%m%d')}.xlsx"
    workbook.save(excel_output_file)

    # Create the table as a formatted string
    table_data = []
    header = [''] + attributes
    table_data.append(header)
    for i, paragraph in enumerate(paragraphs):
        row_values = [paragraph.strip()] + [table_sheet.cell(row=i+2, column=j+2).value or '' for j in range(len(attributes))]
        table_data.append(row_values)

    # Format the table data
    col_widths = [max(len(row[i]) for row in table_data) for i in range(len(header))]
    table_formatted = ''
    for row in table_data:
        row_formatted = ''
        for i, col_width in enumerate(col_widths):
            row_formatted += row[i].ljust(col_width + 2)
        table_formatted += row_formatted.rstrip() + '\n'

    # Save the table as a text file
    text_output_file = f"{output_name}_{date.today().strftime('%Y%m%d')}.txt"
    with open(text_output_file, 'w') as f:
        f.write(table_formatted)

    # Generate a graph based on the data
    plt.figure(figsize=(10, 6))
    x = list(range(len(paragraphs)))
    for i, attr in enumerate(attributes):
        y = [table_sheet.cell(row=j+2, column=i+2).value or 0 for j in range(len(paragraphs))]
        plt.plot(x, y, label=attr)
    plt.xlabel('Paragraphs')
    plt.ylabel('Attribute Values')
    plt.title('Graph of Attribute Values')
    plt.legend()
    plt.xticks(x, paragraphs, rotation=45)
    plt.subplots_adjust(bottom=0.25, top=0.90)  # Adjust bottom and top margins

    # Save the graph as an image file
    graph_output_file = f"{output_name}_{date.today().strftime('%Y%m%d')}.png"
    plt.savefig(graph_output_file)
    plt.close()

    return text_output_file, excel_output_file, graph_output_file


# Example usage
output_files = create_table("paragraphs.txt", "attributes.txt", "output")
print("Text file:", output_files[0])
print("Excel file:", output_files[1])
print("Graph file:", output_files[2])
