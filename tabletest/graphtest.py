import openpyxl
import matplotlib.pyplot as plt

def create_table(paragraph_file, attribute_file, output_file):
    # Read paragraphs from the first file
    with open(paragraph_file, 'r') as f:
        paragraphs = f.readlines()

    # Read attributes from the second file
    with open(attribute_file, 'r') as f:
        attributes = [attr.strip() for attr in f.readlines()]

    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the attributes as the first row
    for i, attr in enumerate(attributes):
        sheet.cell(row=1, column=i+2).value = attr

    # Process each paragraph and extract attribute values
    for i, paragraph in enumerate(paragraphs):
        attr_values = {}
        for attr in attributes:
            attr_key = attr + ':'
            if attr_key in paragraph:
                start_index = paragraph.index(attr_key) + len(attr_key)
                end_index = paragraph.find(' ', start_index)
                if end_index == -1:
                    end_index = None
                attr_value = paragraph[start_index:end_index].strip()
                attr_values[attr] = attr_value

        # Write attribute values as a new row in the sheet
        sheet.cell(row=i+2, column=1).value = paragraph.strip()
        for j, attr in enumerate(attributes):
            attr_value = attr_values.get(attr, '')
            sheet.cell(row=i+2, column=j+2).value = attr_value

    # Save the workbook to the output file
    workbook.save(output_file)

    # Generate a graph based on the data in the sheet
    data = sheet.iter_rows(values_only=True)
    next(data)  # Skip the header row
    data = list(data)
    x = [row[0] for row in data]  # X-axis values (paragraphs)
    y = [row[1:] for row in data]  # Y-axis values (attribute values)

    plt.figure(figsize=(10, 6))
    for i, attr in enumerate(attributes):
        plt.plot(x, [row[i] for row in y], label=attr)
    plt.xlabel('Paragraphs')
    plt.ylabel('Attribute Values')
    plt.title('Graph of Attribute Values')
    plt.legend()
    plt.xticks(rotation=45)
    plt.subplots_adjust(bottom=0.25, top=0.90)  # Adjust bottom and top margins

    # Save the graph as an image file
    graph_output_file = output_file.replace('.xlsx', '.png')
    plt.savefig(graph_output_file)
    plt.close()


# Example usage
create_table("paragraphs.txt", "attributes.txt", "output.xlsx")
